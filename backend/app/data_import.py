"""Safe, previewable bulk data import - customers, products, opening GL
trial balances, and historical invoices. The `products` entity has the
fullest feature set (duplicate/update policy, warehouse-aware opening
stock, rollback) since it backs the Product Import Wizard; the other three
entities are intentionally left exactly as they were before that wizard
was added, so nothing about their behavior changes here.
"""
import io
import json
import uuid
from datetime import datetime, timezone

from fastapi import APIRouter, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from pydantic import BaseModel
from sqlalchemy import text
from sqlalchemy.orm import Session as OrmSession

from app.accounting.integrity import calculate_payment_status, money
from app.accounting.posting import post_balanced_voucher, delete_source_voucher
from app.accounting.posting import _ensure_schema as _ensure_posting_schema
from app.database import SessionLocal, engine
from app.financial_policy import financial_policy_values
from app.warehouses import DEFAULT_WAREHOUSE_NAME, apply_warehouse_delta
from app.warehouses import _ensure_default_warehouse as ensure_default_warehouse
from app.company_scope import current_company_id

router = APIRouter(prefix="/api/data-import", tags=["Safe Data Import"])
MAX_BYTES = 25 * 1024 * 1024
MAX_ROWS = 12000
ENTITY_FIELDS = {
    "customers": {
        "name": ("name", "نام", "نام طرف حساب", "نام طرف‌حساب"),
        "phone": ("phone", "تلفن", "موبایل"),
        "email": ("email", "ایمیل"),
        "address": ("address", "آدرس"),
        "city": ("city", "شهر"),
        "national_id": ("national_id", "شناسه ملی", "کد ملی"),
        "economic_code": ("economic_code", "کد اقتصادی"),
        "contact_person": ("contact_person", "شخص رابط"),
        "customer_type": ("customer_type", "نوع طرف حساب", "نوع طرف‌حساب"),
        "opening_balance": ("opening_balance", "مانده افتتاحیه", "مانده اول دوره"),
        "credit_limit": ("credit_limit", "سقف اعتبار"),
        "notes": ("notes", "یادداشت", "توضیحات"),
    },
    "products": {
        "name": ("name", "نام", "نام کالا", "ProductName", "GoodsName"),
        "code": ("code", "کد", "کد کالا", "ProductCode", "GoodsCode"),
        "barcode": ("barcode", "بارکد", "BarCode", "BarCode1"),
        "barcode2": ("barcode2", "بارکد دوم", "BarCode2"),
        "barcode3": ("barcode3", "بارکد سوم", "BarCode3"),
        "sku": ("sku", "SKU", "StockCode", "شناسه کالا"),
        "brand": ("brand", "برند"),
        "unit": ("unit", "واحد", "UnitName", "Unit"),
        "buy_price": ("buy_price", "قیمت خرید", "PurchasePrice", "BuyPrice"),
        "sell_price": ("sell_price", "قیمت فروش", "SalesPrice", "SellPrice"),
        "stock": ("stock", "موجودی", "موجودی اولیه", "OpeningStock", "FirstAmount", "InitialStock"),
        "min_stock": ("min_stock", "حداقل موجودی"),
        "count_in_pack": ("count_in_pack", "تعداد در بسته", "CountInPack", "PackQuantity"),
        "main_category": ("main_category", "گروه اصلی", "Group", "GroupName", "Category", "CategoryName", "گروه", "دسته‌بندی"),
        "sub_category": ("sub_category", "گروه فرعی"),
        "description": ("description", "توضیحات", "Desc", "Description"),
    },
    "gl_trial_balance": {
        "account_code": ("account_code", "کد حساب", "کد"),
        "account_name": ("account_name", "نام حساب", "شرح حساب"),
        "debit": ("debit", "بدهکار"),
        "credit": ("credit", "بستانکار"),
        "description": ("description", "شرح", "توضیحات"),
    },
    "invoices": {
        "invoice_no": ("invoice_no", "شماره فاکتور", "شماره سند"),
        "date": ("date", "تاریخ", "تاریخ فاکتور"),
        "invoice_type": ("invoice_type", "نوع فاکتور", "نوع سند"),
        "customer_name": ("customer_name", "نام طرف‌حساب", "نام مشتری"),
        "product": ("product", "کالا", "نام یا کد کالا", "نام کالا", "کد کالا"),
        "quantity": ("quantity", "تعداد"),
        "unit_price": ("unit_price", "قیمت واحد", "قیمت فروش واحد"),
        "discount": ("discount", "تخفیف"),
        "paid_amount": ("paid_amount", "مبلغ پرداخت‌شده", "مبلغ دریافتی"),
    },
}
NUMERIC_FIELDS = {
    "opening_balance", "credit_limit", "buy_price", "sell_price", "min_stock",
    "debit", "credit", "quantity", "unit_price", "discount", "paid_amount",
    "stock", "count_in_pack",
}
# Negative values are always rejected for these - "stock" is deliberately
# NOT here: for the products entity a negative opening stock is
# policy-controlled (see negative_stock_policy) rather than an unconditional
# error, since some real-world migrations legitimately start from a known
# shortfall.
ALWAYS_NON_NEGATIVE_FIELDS = {
    "credit_limit", "buy_price", "sell_price", "min_stock", "count_in_pack",
    "debit", "credit", "quantity", "unit_price", "discount", "paid_amount",
}
IDENTIFIER_FIELDS = {"code", "sku", "barcode", "barcode2", "barcode3"}
REQUIRED_FIELD = {
    "customers": "name",
    "products": "name",
    "gl_trial_balance": "account_code",
    "invoices": "date",
}
PRODUCT_MATCH_KEY_FIELDS = {
    "code": ("code",),
    "sku": ("sku",),
    "barcode": ("barcode",),
    "code_or_barcode": ("code", "barcode"),
}
PRODUCT_UPDATABLE_FIELDS = {
    # "code" and "stock" are deliberately excluded: code is the match key
    # itself, and stock follows the same rule as main.py's update_product -
    # once a product exists, on-hand quantity only ever changes through an
    # explicit stock movement, never a silent bulk-update overwrite.
    "barcode", "barcode2", "barcode3", "sku", "brand", "unit", "buy_price",
    "sell_price", "min_stock", "main_category", "sub_category",
    "description", "count_in_pack",
}
# Historical invoice import intentionally supports only sale/return_sale, and
# intentionally never touches Product.stock: current stock is expected to be
# seeded separately via the "products" import, so replaying historical stock
# movements here would double-count inventory. See WINDOWS docs for the full
# migration order and this tradeoff.
SUPPORTED_HISTORICAL_INVOICE_TYPES = {"sale", "return_sale"}
# Preview/inspect/history are readable by anyone who can legitimately handle
# inventory data; apply/rollback stay strictly administrator-only. Only the
# "products" entity gets the wider tier - customers/gl_trial_balance/invoices
# keep the exact admin-only behavior they always had.
PREVIEW_ROLES = {"admin", "accountant", "warehouse"}


def _now():
    return datetime.now(timezone.utc).isoformat()


def _actor_id(auth):
    try:
        return int(auth["sub"])
    except (KeyError, TypeError, ValueError):
        raise HTTPException(status_code=401, detail="Invalid authentication context")


def _require_admin(request):
    auth = getattr(request.state, "auth", {})
    if auth.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Administrator access required")
    return _actor_id(auth)


def _require_preview_role(request, entity=None):
    auth = getattr(request.state, "auth", {})
    allowed = PREVIEW_ROLES if entity == "products" else {"admin"}
    if auth.get("role") not in allowed:
        raise HTTPException(status_code=403, detail=f"Access requires one of: {', '.join(sorted(allowed))}")
    return _actor_id(auth)


def _ensure_schema(conn):
    conn.execute(text("""
        CREATE TABLE IF NOT EXISTS data_import_batches (
            id VARCHAR PRIMARY KEY,
            entity_type VARCHAR NOT NULL,
            file_name VARCHAR NOT NULL,
            status VARCHAR NOT NULL,
            total_rows INTEGER NOT NULL,
            valid_rows INTEGER NOT NULL,
            error_rows INTEGER NOT NULL,
            duplicate_rows INTEGER NOT NULL,
            payload_json TEXT NOT NULL,
            result_json TEXT DEFAULT '',
            created_by INTEGER NOT NULL,
            created_at VARCHAR NOT NULL,
            applied_at VARCHAR,
            FOREIGN KEY(created_by) REFERENCES users(id)
        )
    """))
    columns = {row[1] for row in conn.execute(text("PRAGMA table_info(data_import_batches)")).fetchall()}
    if "meta_json" not in columns:
        conn.execute(text("ALTER TABLE data_import_batches ADD COLUMN meta_json TEXT DEFAULT '{}'"))
    conn.execute(text("""
        CREATE TABLE IF NOT EXISTS data_import_product_links (
            batch_id VARCHAR NOT NULL,
            product_id INTEGER NOT NULL,
            created_at VARCHAR NOT NULL
        )
    """))
    conn.execute(text("""
        CREATE TABLE IF NOT EXISTS data_import_mapping_profiles (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            entity_type VARCHAR NOT NULL,
            name VARCHAR NOT NULL,
            mapping_json TEXT NOT NULL,
            created_by INTEGER NOT NULL,
            created_at VARCHAR NOT NULL
        )
    """))
    from app.company_scope import ensure_company_id_column
    ensure_company_id_column(conn, "data_import_batches")
    ensure_company_id_column(conn, "data_import_product_links")
    ensure_company_id_column(conn, "data_import_mapping_profiles")


def _clean(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _clean_identifier(value):
    """Like _clean(), but for code/sku/barcode columns specifically: Excel
    can hand back a numeric-looking identifier as a float, including
    scientific notation for long digit strings (e.g. a 13-digit barcode
    rendered as 8.901234568e+12). format(value, "f") always renders the
    full integer with no exponent and no precision loss, unlike str()."""
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return format(value, ".0f")
        return format(value, "f").rstrip("0").rstrip(".")
    return str(value).strip()


def _number(value, field, row_number, errors):
    if value in (None, ""):
        return 0.0
    try:
        result = float(str(value).replace(",", "").replace("٬", ""))
        if field in ALWAYS_NON_NEGATIVE_FIELDS and result < 0:
            errors.append({"row": row_number, "field": field, "message": "Value cannot be negative"})
        return result
    except (TypeError, ValueError):
        errors.append({"row": row_number, "field": field, "message": "Invalid number"})
        return 0.0


def _parse_date(value):
    text_value = _clean(value)[:10]
    try:
        return datetime.strptime(text_value, "%Y-%m-%d").date()
    except ValueError:
        return None


def _mapping(headers, entity):
    aliases = ENTITY_FIELDS[entity]
    normalized = {_clean(value).lower(): index for index, value in enumerate(headers)}
    result = {}
    for field, names in aliases.items():
        for name in names:
            if name.lower() in normalized:
                result[field] = normalized[name.lower()]
                break
    required = REQUIRED_FIELD[entity]
    if required not in result:
        raise HTTPException(status_code=400, detail=f"Required '{required}' column is missing")
    return result


def _resolve_mapping_override(headers, entity, override_json):
    try:
        override = json.loads(override_json)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="Invalid column_map_json: must be a JSON object")
    if not isinstance(override, dict):
        raise HTTPException(status_code=400, detail="Invalid column_map_json: must be a JSON object")
    normalized = {_clean(value).lower(): index for index, value in enumerate(headers)}
    result = {}
    for field, header_name in override.items():
        if field not in ENTITY_FIELDS[entity]:
            continue
        key = _clean(header_name).lower()
        if key in normalized:
            result[field] = normalized[key]
    required = REQUIRED_FIELD[entity]
    if required not in result:
        raise HTTPException(status_code=400, detail=f"Required '{required}' column is missing from column_map_json")
    return result


def _product_match_key(item, match_keys):
    fields = PRODUCT_MATCH_KEY_FIELDS.get(match_keys) or ("code", "barcode")
    for field in fields:
        value = item.get(field)
        if value:
            return (field, str(value).lower())
    return ("name", str(item.get("name") or "").lower())


def _existing_keys(conn, entity, match_keys=None, company_id=None):
    if entity == "customers":
        rows = conn.execute(
            text("SELECT name, phone, national_id FROM customers WHERE company_id=:company_id"),
            {"company_id": company_id},
        ).mappings()
        return {
            ("national", _clean(row["national_id"]).lower()) if _clean(row["national_id"])
            else ("name_phone", _clean(row["name"]).lower(), _clean(row["phone"]).lower())
            for row in rows
        }
    if entity == "products":
        rows = conn.execute(
            text("SELECT name, code, barcode, sku FROM products WHERE company_id=:company_id"),
            {"company_id": company_id},
        ).mappings()
        if match_keys:
            result = set()
            for row in rows:
                result.add(_product_match_key({k: _clean(v) for k, v in dict(row).items()}, match_keys))
            return result
        return {
            ("code", _clean(row["code"]).lower()) if _clean(row["code"])
            else ("barcode", _clean(row["barcode"]).lower()) if _clean(row["barcode"])
            else ("name", _clean(row["name"]).lower())
            for row in rows
        }
    # gl_trial_balance / invoices: rows are lines, not standalone records -
    # repeats are expected and legitimate (e.g. two lines on the same
    # account, or two lines in the same invoice), so duplicate detection
    # is intentionally skipped via _row_key below.
    return set()


def _existing_product_ids_by_key(conn, match_keys, company_id):
    rows = conn.execute(
        text("SELECT id, name, code, barcode, sku FROM products WHERE company_id=:company_id"),
        {"company_id": company_id},
    ).mappings()
    result = {}
    for row in rows:
        item = {k: _clean(v) for k, v in dict(row).items()}
        result[_product_match_key(item, match_keys)] = row["id"]
    return result


def _row_key(entity, row, row_number=None, match_keys=None):
    if entity == "customers":
        return (
            ("national", row["national_id"].lower())
            if row.get("national_id")
            else ("name_phone", row["name"].lower(), row.get("phone", "").lower())
        )
    if entity == "products":
        if match_keys:
            return _product_match_key(row, match_keys)
        return (
            ("code", row["code"].lower()) if row.get("code")
            else ("barcode", row["barcode"].lower()) if row.get("barcode")
            else ("name", row["name"].lower())
        )
    return ("row", row_number)


def _customer_lookup(conn, company_id):
    rows = conn.execute(
        text("SELECT id, name FROM customers WHERE company_id=:company_id"),
        {"company_id": company_id},
    ).mappings()
    return {_clean(row["name"]).lower(): row["id"] for row in rows if _clean(row["name"])}


def _product_lookup(conn, company_id):
    rows = conn.execute(
        text("SELECT id, name, code, buy_price FROM products WHERE company_id=:company_id"),
        {"company_id": company_id},
    ).mappings()
    lookup = {}
    for row in rows:
        buy_price = float(row["buy_price"] or 0)
        if _clean(row["code"]):
            lookup[_clean(row["code"]).lower()] = (row["id"], buy_price)
        if _clean(row["name"]):
            lookup.setdefault(_clean(row["name"]).lower(), (row["id"], buy_price))
    return lookup


@router.get("/template/{entity}")
def download_template(entity: str, request: Request, language: str = "en"):
    _require_preview_role(request, entity=entity)
    if entity not in ENTITY_FIELDS:
        raise HTTPException(status_code=404, detail="Unsupported import entity")
    fa = language == "fa"
    headers = {
        "customers": (
            ["نام طرف‌حساب", "تلفن", "ایمیل", "آدرس", "شهر", "شناسه ملی",
             "کد اقتصادی", "شخص رابط", "نوع طرف‌حساب", "مانده افتتاحیه",
             "سقف اعتبار", "یادداشت"]
            if fa else
            ["name", "phone", "email", "address", "city", "national_id",
             "economic_code", "contact_person", "customer_type", "opening_balance",
             "credit_limit", "notes"]
        ),
        "products": (
            ["نام کالا", "کد کالا", "بارکد", "بارکد دوم", "بارکد سوم", "SKU", "برند", "واحد",
             "قیمت خرید", "قیمت فروش", "موجودی اولیه", "حداقل موجودی", "تعداد در بسته",
             "گروه اصلی", "گروه فرعی", "توضیحات"]
            if fa else
            ["name", "code", "barcode", "barcode2", "barcode3", "sku", "brand", "unit", "buy_price",
             "sell_price", "stock", "min_stock", "count_in_pack", "main_category", "sub_category", "description"]
        ),
        "gl_trial_balance": (
            ["کد حساب", "نام حساب", "بدهکار", "بستانکار", "شرح"]
            if fa else
            ["account_code", "account_name", "debit", "credit", "description"]
        ),
        "invoices": (
            ["شماره فاکتور", "تاریخ (YYYY-MM-DD)", "نوع فاکتور (sale/return_sale)",
             "نام طرف‌حساب", "نام یا کد کالا", "تعداد", "قیمت واحد", "تخفیف", "مبلغ پرداخت‌شده"]
            if fa else
            ["invoice_no", "date (YYYY-MM-DD)", "invoice_type (sale/return_sale)",
             "customer_name", "product (name or code)", "quantity", "unit_price", "discount", "paid_amount"]
        ),
    }[entity]
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = {
        "customers": "طرف‌حساب‌ها" if fa else "Customers",
        "products": "کالاها" if fa else "Products",
        "gl_trial_balance": "سند افتتاحیه" if fa else "Opening trial balance",
        "invoices": "فاکتورهای تاریخی" if fa else "Historical invoices",
    }[entity]
    sheet.sheet_view.rightToLeft = fa
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F172A")
    if entity == "products":
        # Code/SKU/barcode columns must round-trip through Excel as text -
        # otherwise a leading-zero or long numeric-looking code silently
        # turns into a float (or scientific notation) before the user even
        # gets to typing real data into the template.
        for col_letter in ("B", "C", "D", "E", "F"):
            sheet.column_dimensions[col_letter].number_format = "@"
    sample = {
        "customers": ["نمونه", "09120000000", "sample@example.com", "", "", "", "", "", "customer", 0, 0, ""],
        "products": ["نمونه کالا", "P-001", "", "", "", "", "", "عدد", 0, 0, 0, 0, 1, "", "", ""],
        "gl_trial_balance": ["1101", "صندوق", 1000000, 0, "مانده اول دوره صندوق"],
        "invoices": ["1001", "2025-01-15", "sale", "نمونه مشتری", "P-001", 2, 500000, 0, 1000000],
    }[entity]
    if not fa:
        if entity == "customers":
            sample[0] = "Sample"
        elif entity == "products":
            sample[0] = "Sample product"
            sample[7] = "unit"
        elif entity == "gl_trial_balance":
            sample[1] = "Cash"
            sample[4] = "Opening cash balance"
        elif entity == "invoices":
            sample[3] = "Sample customer"
    sheet.append(sample)
    sheet.freeze_panes = "A2"
    for column in sheet.columns:
        sheet.column_dimensions[column[0].column_letter].width = 22
    guide = workbook.create_sheet("Guide" if not fa else "راهنما")
    guide.sheet_view.rightToLeft = fa
    required_field = REQUIRED_FIELD[entity]
    guide_lines = (
        [
            f"ستون اجباری: {required_field}",
            "بقیه ستون‌ها اختیاری‌اند؛ اگر خالی بمانند، مقدار پیش‌فرض یا خالی ثبت می‌شود.",
            "کد کالا، SKU و بارکد باید به‌صورت متن (Text) نگه داشته شوند تا صفر ابتدایی یا اعداد بزرگ خراب نشوند.",
        ]
        if fa else
        [
            f"Required column: {required_field}",
            "All other columns are optional; blank cells use a default or stay empty.",
            "Keep code/SKU/barcode columns formatted as Text so leading zeros or long digit strings aren't corrupted.",
        ]
    )
    for line in guide_lines:
        guide.append([line])
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"vetrix_{entity}_import_{'fa' if fa else 'en'}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/inspect")
async def inspect_import_file(
    request: Request,
    file: UploadFile = File(...),
    entity: str = Form("products"),
    sheet_name: str = Form(""),
):
    _require_preview_role(request, entity=entity)
    raw = await file.read(MAX_BYTES + 1)
    if len(raw) > MAX_BYTES:
        raise HTTPException(status_code=413, detail=f"File exceeds {MAX_BYTES // (1024 * 1024)} MB limit")
    if not (file.filename or "").lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx or .xls files are accepted")
    if (file.filename or "").lower().endswith(".xls"):
        raise HTTPException(status_code=400, detail="Legacy .xls is not supported - please save as .xlsx and retry")
    try:
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as error:
        raise HTTPException(status_code=400, detail=f"Invalid Excel workbook: {error}")
    row_counts = {}
    for name in workbook.sheetnames:
        count = sum(1 for _ in workbook[name].iter_rows(values_only=True))
        row_counts[name] = max(0, count - 1)

    sheet = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    headers = next(rows_iter, None) or []
    sample_row = next(rows_iter, None) or []
    headers_clean = [_clean(h) for h in headers]

    suggested_mapping = {}
    if entity in ENTITY_FIELDS:
        try:
            suggested_mapping = _mapping(headers, entity)
        except HTTPException:
            suggested_mapping = {}

    return {
        "file_name": file.filename,
        "size_bytes": len(raw),
        "sheet_names": workbook.sheetnames,
        "row_counts": row_counts,
        "active_sheet": sheet.title,
        "headers": headers_clean,
        "sample_row": [_clean(v) for v in sample_row],
        "suggested_mapping": suggested_mapping,
        "entity_fields": list(ENTITY_FIELDS.get(entity, {}).keys()),
    }


@router.get("/mapping-profiles")
def list_mapping_profiles(request: Request, entity: str = "products"):
    _require_preview_role(request, entity=entity)
    with engine.begin() as conn:
        _ensure_schema(conn)
        rows = conn.execute(text("""
            SELECT id, entity_type, name, mapping_json, created_by, created_at
            FROM data_import_mapping_profiles WHERE entity_type=:entity AND company_id=:company_id ORDER BY created_at DESC
        """), {"entity": entity, "company_id": current_company_id(request)}).mappings().all()
        return [dict(row) for row in rows]


class MappingProfileCreate(BaseModel):
    entity_type: str
    name: str
    mapping: dict


@router.post("/mapping-profiles")
def save_mapping_profile(data: MappingProfileCreate, request: Request):
    actor = _require_preview_role(request, entity=data.entity_type)
    if data.entity_type not in ENTITY_FIELDS:
        raise HTTPException(status_code=404, detail="Unsupported import entity")
    if not data.name.strip():
        raise HTTPException(status_code=400, detail="Profile name is required")
    with engine.begin() as conn:
        _ensure_schema(conn)
        conn.execute(text("""
            INSERT INTO data_import_mapping_profiles (entity_type, name, mapping_json, created_by, created_at, company_id)
            VALUES (:entity, :name, :mapping, :actor, :created_at, :company_id)
        """), {
            "entity": data.entity_type, "name": data.name.strip(),
            "mapping": json.dumps(data.mapping, ensure_ascii=False),
            "actor": actor, "created_at": _now(),
            "company_id": current_company_id(request),
        })
        return {"status": "created"}


@router.post("/preview/{entity}")
async def preview_import(
    entity: str,
    request: Request,
    file: UploadFile = File(...),
    opening_date: str = Form(""),
    sheet_name: str = Form(""),
    column_map_json: str = Form(""),
    duplicate_policy: str = Form("skip"),
    match_keys: str = Form("code_or_barcode"),
    negative_stock_policy: str = Form("block"),
    default_warehouse_id: str = Form(""),
    update_fields: str = Form(""),
    allow_blank_overwrite: bool = Form(False),
):
    actor = _require_preview_role(request, entity=entity)
    if entity not in ENTITY_FIELDS:
        raise HTTPException(status_code=404, detail="Unsupported import entity")
    raw = await file.read(MAX_BYTES + 1)
    if len(raw) > MAX_BYTES:
        raise HTTPException(status_code=413, detail=f"File exceeds {MAX_BYTES // (1024 * 1024)} MB limit")
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are accepted")
    try:
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        sheet = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = next(rows)
    except HTTPException:
        raise
    except Exception as error:
        raise HTTPException(status_code=400, detail=f"Invalid Excel workbook: {error}")
    column_map = (
        _resolve_mapping_override(headers, entity, column_map_json)
        if column_map_json else
        _mapping(headers, entity)
    )

    is_products = entity == "products"
    duplicate_policy = duplicate_policy if duplicate_policy in {"skip", "update"} else "skip"
    match_keys = match_keys if match_keys in PRODUCT_MATCH_KEY_FIELDS else "code_or_barcode"
    negative_stock_policy = negative_stock_policy if negative_stock_policy in {"block", "warn"} else "block"
    update_fields_list = [f.strip() for f in update_fields.split(",") if f.strip()] if update_fields else []
    warehouse_id_int = None
    if is_products and default_warehouse_id:
        try:
            warehouse_id_int = int(default_warehouse_id)
        except ValueError:
            warehouse_id_int = None

    parsed, errors, warnings = [], [], []
    seen = set()
    with engine.begin() as conn:
        _ensure_schema(conn)
        existing = _existing_keys(conn, entity, match_keys=match_keys if is_products else None, company_id=current_company_id(request))
        chart_codes = set()
        customers_map = {}
        products_map = {}
        if entity == "gl_trial_balance":
            _ensure_posting_schema(conn)
            chart_codes = {row[0] for row in conn.execute(text("SELECT code FROM chart_accounts WHERE company_id=:company_id"), {"company_id": current_company_id(request)}).fetchall()}
        if entity == "invoices":
            customers_map = _customer_lookup(conn, current_company_id(request))
            products_map = _product_lookup(conn, current_company_id(request))
        total_debit = 0.0
        total_credit = 0.0
        for row_number, values in enumerate(rows, start=2):
            if row_number - 1 > MAX_ROWS:
                raise HTTPException(status_code=400, detail=f"Maximum {MAX_ROWS} data rows allowed")
            if not any(value not in (None, "") for value in values):
                continue
            item = {}
            row_errors = []
            row_warnings = []
            for field in ENTITY_FIELDS[entity]:
                value = values[column_map[field]] if field in column_map and column_map[field] < len(values) else ""
                if field in NUMERIC_FIELDS:
                    item[field] = _number(value, field, row_number, row_errors)
                elif field in IDENTIFIER_FIELDS:
                    item[field] = _clean_identifier(value)
                else:
                    item[field] = _clean(value)
            required = REQUIRED_FIELD[entity]
            if not item.get(required):
                row_errors.append({"row": row_number, "field": required, "message": f"{required} is required"})

            if entity == "customers":
                if item["customer_type"] not in {"", "customer", "supplier", "both"}:
                    row_errors.append({"row": row_number, "field": "customer_type", "message": "Use customer, supplier, or both"})
                item["customer_type"] = item.get("customer_type") or "customer"
            elif entity == "products":
                item["unit"] = item.get("unit") or "عدد"
                if not item.get("sell_price"):
                    row_warnings.append({"row": row_number, "field": "sell_price", "message": "Sell price is empty or zero"})
                if item.get("stock", 0) < 0:
                    if negative_stock_policy == "block":
                        row_errors.append({"row": row_number, "field": "stock", "message": "Negative stock is not allowed"})
                    else:
                        row_warnings.append({"row": row_number, "field": "stock", "message": "Negative stock accepted (policy: warn)"})
            elif entity == "gl_trial_balance":
                debit, credit = item.get("debit") or 0, item.get("credit") or 0
                if debit and credit:
                    row_errors.append({"row": row_number, "field": "debit", "message": "Only one of debit/credit is allowed per row"})
                elif not debit and not credit:
                    row_errors.append({"row": row_number, "field": "debit", "message": "A debit or credit amount is required"})
                if item.get("account_code") and item["account_code"] not in chart_codes:
                    row_errors.append({"row": row_number, "field": "account_code", "message": "Unknown account code"})
                total_debit += debit
                total_credit += credit
            elif entity == "invoices":
                item["invoice_type"] = (item.get("invoice_type") or "sale").strip().lower()
                if item["invoice_type"] not in SUPPORTED_HISTORICAL_INVOICE_TYPES:
                    row_errors.append({"row": row_number, "field": "invoice_type", "message": "Only sale or return_sale is supported for historical import"})
                if item.get("date") and not _parse_date(item["date"]):
                    row_errors.append({"row": row_number, "field": "date", "message": "Invalid date, use YYYY-MM-DD"})
                if not item.get("customer_name"):
                    row_errors.append({"row": row_number, "field": "customer_name", "message": "Customer name is required"})
                elif item["customer_name"].lower() not in customers_map:
                    row_errors.append({"row": row_number, "field": "customer_name", "message": "Customer not found - import customers first"})
                if not item.get("product"):
                    row_errors.append({"row": row_number, "field": "product", "message": "Product is required"})
                elif item["product"].lower() not in products_map:
                    row_errors.append({"row": row_number, "field": "product", "message": "Product not found - import products first"})
                if not item.get("quantity") or item["quantity"] <= 0:
                    row_errors.append({"row": row_number, "field": "quantity", "message": "Quantity must be greater than zero"})

            key = _row_key(entity, item, row_number, match_keys=match_keys if is_products else None)
            is_dup_key = bool(key) and (key in seen or key in existing) and entity in {"customers", "products"}
            updatable = is_dup_key and is_products and duplicate_policy == "update"
            duplicate = is_dup_key and not updatable
            if is_dup_key:
                warnings.append({
                    "row": row_number, "field": "duplicate",
                    "message": "Duplicate row will be updated" if updatable else "Duplicate row will be skipped",
                })
            seen.add(key)
            parsed.append({
                "row": row_number, "data": item, "errors": row_errors,
                "duplicate": duplicate, "updatable": updatable, "warnings": row_warnings,
            })
            errors.extend(row_errors)
            warnings.extend(row_warnings)
        if entity == "gl_trial_balance" and parsed:
            rounded_debit = round(total_debit, 2)
            rounded_credit = round(total_credit, 2)
            if rounded_debit != rounded_credit:
                warnings.append({
                    "row": 0, "field": "balance",
                    "message": f"Trial balance is not balanced: debit={rounded_debit}, credit={rounded_credit}. Fix before applying.",
                })
        batch_id = str(uuid.uuid4())
        valid = sum(not row["errors"] and not row["duplicate"] for row in parsed)
        updatable_rows = sum(row["updatable"] for row in parsed)
        balanced = entity != "gl_trial_balance" or round(total_debit, 2) == round(total_credit, 2)
        meta = {
            "opening_date": opening_date or datetime.utcnow().date().isoformat(),
            "duplicate_policy": duplicate_policy,
            "match_keys": match_keys,
            "negative_stock_policy": negative_stock_policy,
            "default_warehouse_id": warehouse_id_int,
            "update_fields": update_fields_list,
            "allow_blank_overwrite": bool(allow_blank_overwrite),
        }
        conn.execute(text("""
            INSERT INTO data_import_batches
              (id, entity_type, file_name, status, total_rows, valid_rows,
               error_rows, duplicate_rows, payload_json, meta_json, created_by, created_at, company_id)
            VALUES
              (:id, :entity, :file_name, 'previewed', :total, :valid,
               :errors, :duplicates, :payload, :meta, :actor, :created_at, :company_id)
        """), {
            "id": batch_id, "entity": entity, "file_name": file.filename or "import.xlsx",
            "total": len(parsed), "valid": valid,
            "errors": sum(bool(row["errors"]) for row in parsed),
            "duplicates": sum(row["duplicate"] for row in parsed),
            "payload": json.dumps(parsed, ensure_ascii=False),
            "meta": json.dumps(meta, ensure_ascii=False),
            "actor": actor, "created_at": _now(),
            "company_id": current_company_id(request),
        })
    return {
        "batch_id": batch_id, "entity": entity, "total_rows": len(parsed),
        "valid_rows": valid, "error_rows": sum(bool(row["errors"]) for row in parsed),
        "duplicate_rows": sum(row["duplicate"] for row in parsed),
        "updatable_rows": updatable_rows,
        # Full per-row detail (up to MAX_ROWS) so the wizard's preview table
        # can filter/sort/paginate client-side over the whole file, not just
        # a truncated slice - each row already carries its own errors/
        # warnings sublist, so capping the flat summary arrays below doesn't
        # lose any per-row detail.
        "errors": errors[:500], "warnings": warnings[:500], "preview": parsed,
        "can_apply": valid > 0 and not errors and balanced,
    }


def _post_customer_opening(conn, customer_id, name, balance, policy, company_id):
    amount = float(money(abs(balance), policy["decimal_places"], policy["rounding_mode"]))
    if not amount:
        return
    description = f"Imported opening balance: {name}"
    # Both sides are required: the GL voucher (chart_accounts/trial balance)
    # AND this customer's own accounting_entries subledger row - the latter
    # is what customer_balance()/the customer ledger and aging report
    # actually read. Posting only the GL side (as this used to) left every
    # imported customer's own balance silently at zero.
    if balance > 0:
        _insert_accounting_entry(conn, customer_id, "opening_balance", customer_id, description, amount, 0, datetime.utcnow(), company_id)
        lines = [{"account_code": "1103", "debit": amount}, {"account_code": "3101", "credit": amount}]
    else:
        _insert_accounting_entry(conn, customer_id, "opening_balance", customer_id, description, 0, amount, datetime.utcnow(), company_id)
        lines = [{"account_code": "3101", "debit": amount}, {"account_code": "2101", "credit": amount}]
    post_balanced_voucher("customer_opening", customer_id, description, lines, company_id, connection=conn)
    _rebuild_customer_balance(conn, customer_id)


def _post_product_opening(conn, product_id, name, stock, buy_price, policy, company_id):
    amount = float(money(stock * buy_price, policy["decimal_places"], policy["rounding_mode"]))
    if not amount:
        return
    description = f"Imported opening inventory: {name}"
    post_balanced_voucher("product_opening", product_id, description, [
        {"account_code": "1201", "debit": amount},
        {"account_code": "3101", "credit": amount},
    ], company_id, connection=conn)


def _insert_opening_stock_movement(conn, product_id, product_name, quantity, batch_id, warehouse_id):
    warehouse_name = DEFAULT_WAREHOUSE_NAME
    if warehouse_id:
        row = conn.execute(text("SELECT name FROM warehouses WHERE id=:id"), {"id": warehouse_id}).first()
        if row:
            warehouse_name = row[0]
    now = datetime.utcnow()
    conn.execute(text("""
        INSERT INTO stock_movements
          (warehouse, product_id, product_name, quantity, movement_type, movement_date, note, created_at, warehouse_id)
        VALUES
          (:warehouse, :product_id, :product_name, :quantity, 'opening_import', :movement_date, :note, :created_at, :warehouse_id)
    """), {
        "warehouse": warehouse_name, "product_id": product_id, "product_name": product_name,
        "quantity": quantity, "movement_date": now.date().isoformat(),
        "note": f"Import batch {batch_id}", "created_at": now.isoformat(), "warehouse_id": warehouse_id,
    })


def _apply_product_update(conn, product_id, item, update_fields, allow_blank_overwrite):
    if not update_fields:
        return
    set_clauses = []
    params = {"id": product_id}
    for field in update_fields:
        if field not in PRODUCT_UPDATABLE_FIELDS:
            continue
        value = item.get(field)
        is_blank = value in (None, "", 0, 0.0)
        if is_blank and not allow_blank_overwrite:
            continue
        set_clauses.append(f"{field} = :{field}")
        params[field] = value
    if not set_clauses:
        return
    conn.execute(text(f"UPDATE products SET {', '.join(set_clauses)} WHERE id = :id"), params)


def _insert_accounting_entry(conn, customer_id, source_type, source_id, description, debit, credit, created_at_dt, company_id):
    entry_type = "debit" if debit >= credit else "credit"
    conn.execute(text("""
        INSERT INTO accounting_entries
          (customer_id, source_type, source_id, entry_type, description, debit, credit, balance_after, created_at, company_id)
        VALUES
          (:customer_id, :source_type, :source_id, :entry_type, :description, :debit, :credit, 0, :created_at, :company_id)
    """), {
        "customer_id": customer_id, "source_type": source_type, "source_id": source_id,
        "entry_type": entry_type, "description": description, "debit": debit, "credit": credit,
        "created_at": created_at_dt, "company_id": company_id,
    })


def _rebuild_customer_balance(conn, customer_id):
    rows = conn.execute(text("""
        SELECT id, debit, credit FROM accounting_entries
        WHERE customer_id=:cid ORDER BY created_at ASC, id ASC
    """), {"cid": customer_id}).fetchall()
    balance = 0.0
    for entry_id, debit, credit in rows:
        balance += float(debit or 0) - float(credit or 0)
        conn.execute(text("UPDATE accounting_entries SET balance_after=:b WHERE id=:id"), {"b": balance, "id": entry_id})


def _post_historical_invoice(conn, invoice_id, customer_id, invoice_type, total_amount, paid_amount, invoice_dt, policy, company_id):
    amount = float(money(total_amount, policy["decimal_places"], policy["rounding_mode"]))
    if amount <= 0:
        return
    voucher_date = invoice_dt.date().isoformat()
    description = f"Imported historical invoice #{invoice_id}"
    is_return = invoice_type == "return_sale"
    debit = 0 if is_return else amount
    credit = amount if is_return else 0
    _insert_accounting_entry(conn, customer_id, "invoice", invoice_id, description, debit, credit, invoice_dt, company_id)
    if is_return:
        lines = [
            {"account_code": "4102", "debit": amount, "description": description},
            {"account_code": "1103", "credit": amount, "description": description},
        ]
    else:
        lines = [
            {"account_code": "1103", "debit": amount, "description": description},
            {"account_code": "4101", "credit": amount, "description": description},
        ]
    post_balanced_voucher("invoice", invoice_id, description, lines, company_id, voucher_date=voucher_date, connection=conn)

    paid = float(money(min(max(paid_amount, 0), amount), policy["decimal_places"], policy["rounding_mode"]))
    if paid > 0:
        receipt_description = f"Imported settlement for historical invoice #{invoice_id}"
        if is_return:
            _insert_accounting_entry(conn, customer_id, "payment", invoice_id, receipt_description, paid, 0, invoice_dt, company_id)
            post_balanced_voucher("payment", invoice_id, receipt_description, [
                {"account_code": "1103", "debit": paid, "description": receipt_description},
                {"account_code": "1102", "credit": paid, "description": receipt_description},
            ], company_id, voucher_date=voucher_date, connection=conn)
        else:
            _insert_accounting_entry(conn, customer_id, "receipt", invoice_id, receipt_description, 0, paid, invoice_dt, company_id)
            post_balanced_voucher("receipt", invoice_id, receipt_description, [
                {"account_code": "1102", "debit": paid, "description": receipt_description},
                {"account_code": "1103", "credit": paid, "description": receipt_description},
            ], company_id, voucher_date=voucher_date, connection=conn)
    _rebuild_customer_balance(conn, customer_id)


@router.post("/apply/{batch_id}")
def apply_import(batch_id: str, request: Request):
    actor = _require_admin(request)
    company_id = current_company_id(request)
    with engine.begin() as conn:
        _ensure_schema(conn)
        batch = conn.execute(text(
            "SELECT * FROM data_import_batches WHERE id=:id AND company_id=:company_id"
        ), {"id": batch_id, "company_id": company_id}).mappings().first()
        if not batch:
            raise HTTPException(status_code=404, detail="Import batch not found")
        if batch["created_by"] != actor:
            raise HTTPException(status_code=403, detail="Only the preview creator can apply this batch")
        if batch["status"] != "previewed":
            raise HTTPException(status_code=409, detail="Import batch has already been applied")
        rows = json.loads(batch["payload_json"])
        if any(row["errors"] for row in rows):
            raise HTTPException(status_code=409, detail="Resolve validation errors before applying")
        meta = json.loads(batch["meta_json"] or "{}")
        policy = financial_policy_values(conn, company_id)
        entity = batch["entity_type"]
        inserted = skipped = updated = 0

        if entity == "customers":
            existing = _existing_keys(conn, entity, company_id=company_id)
            for row in rows:
                item = row["data"]
                if row["duplicate"] or _row_key(entity, item) in existing:
                    skipped += 1
                    continue
                # portal_access_enabled/pricing_group/etc. are NOT NULL
                # columns with only an ORM-level (not DB-level) default,
                # so a fresh database (Base.metadata.create_all, e.g. a
                # brand-new install) rejects an INSERT that omits them -
                # unlike an older DB where they were added via an ALTER
                # TABLE ... DEFAULT. Set them explicitly to stay correct
                # on both.
                result = conn.execute(text("""
                    INSERT INTO customers
                      (name, phone, email, address, city, national_id, economic_code,
                       contact_person, customer_type, opening_balance, credit_limit, notes, created_at,
                       portal_access_enabled, portal_token_generation, pricing_group,
                       supplier_portal_access_enabled, supplier_portal_token_generation, company_id)
                    VALUES
                      (:name, :phone, :email, :address, :city, :national_id, :economic_code,
                       :contact_person, :customer_type, :opening_balance, :credit_limit, :notes, :created_at,
                       0, 0, 'retail', 0, 0, :company_id)
                """), {**item, "created_at": datetime.utcnow(), "company_id": company_id})
                _post_customer_opening(conn, result.lastrowid, item["name"], item["opening_balance"], policy, company_id)
                existing.add(_row_key(entity, item))
                inserted += 1

        elif entity == "products":
            duplicate_policy = meta.get("duplicate_policy", "skip")
            match_keys = meta.get("match_keys", "code_or_barcode")
            update_fields = meta.get("update_fields") or []
            allow_blank_overwrite = bool(meta.get("allow_blank_overwrite"))
            warehouse_id = meta.get("default_warehouse_id")

            if warehouse_id:
                # Guarantee the default ("Main") warehouse row exists via an
                # independent session BEFORE the main transaction touches
                # apply_warehouse_delta - its lazy-create path calls db.commit()
                # internally when the row is missing, which would otherwise
                # prematurely commit this batch's own transaction.
                standalone = SessionLocal()
                try:
                    ensure_default_warehouse(standalone, company_id)
                finally:
                    standalone.close()

            existing_by_key = _existing_product_ids_by_key(conn, match_keys, company_id)
            orm_session = OrmSession(bind=conn) if warehouse_id else None

            for row in rows:
                item = row["data"]
                key = _row_key(entity, item, match_keys=match_keys)
                existing_id = existing_by_key.get(key)

                if existing_id and duplicate_policy != "update":
                    skipped += 1
                    continue

                if existing_id and duplicate_policy == "update":
                    _apply_product_update(conn, existing_id, item, update_fields, allow_blank_overwrite)
                    updated += 1
                    continue

                result = conn.execute(text("""
                    INSERT INTO products
                      (name, code, barcode, barcode2, barcode3, sku, brand, unit, buy_price, sell_price,
                       price, stock, min_stock, count_in_pack, main_category, sub_category, image, description, company_id)
                    VALUES
                      (:name, :code, :barcode, :barcode2, :barcode3, :sku, :brand, :unit, :buy_price, :sell_price,
                       :sell_price, :stock, :min_stock, :count_in_pack, :main_category, :sub_category, '', :description, :company_id)
                """), {**item, "company_id": company_id})
                product_id = result.lastrowid
                _post_product_opening(conn, product_id, item["name"], item["stock"], item["buy_price"], policy, company_id)

                if item.get("stock"):
                    _insert_opening_stock_movement(conn, product_id, item["name"], item["stock"], batch_id, warehouse_id)
                    if orm_session is not None:
                        apply_warehouse_delta(orm_session, warehouse_id, product_id, item["stock"], company_id)

                conn.execute(text("""
                    INSERT INTO data_import_product_links (batch_id, product_id, created_at)
                    VALUES (:batch_id, :product_id, :created_at)
                """), {"batch_id": batch_id, "product_id": product_id, "created_at": _now()})

                existing_by_key[key] = product_id
                inserted += 1

            if orm_session is not None:
                orm_session.flush()

        elif entity == "gl_trial_balance":
            _ensure_posting_schema(conn)
            lines = []
            for row in rows:
                item = row["data"]
                debit, credit = item.get("debit") or 0, item.get("credit") or 0
                if not debit and not credit:
                    continue
                lines.append({
                    "account_code": item["account_code"],
                    "debit": debit, "credit": credit,
                    "description": item.get("description") or item.get("account_name") or "Imported opening balance",
                })
            try:
                post_balanced_voucher(
                    f"gl_opening_{batch_id[:12]}", None,
                    "Imported opening trial balance", lines, company_id,
                    voucher_date=meta.get("opening_date"), connection=conn,
                )
            except ValueError as error:
                raise HTTPException(status_code=400, detail=str(error))
            inserted = len(lines)

        elif entity == "invoices":
            customers_map = _customer_lookup(conn, company_id)
            products_map = _product_lookup(conn, company_id)
            groups = {}
            for row in rows:
                item = row["data"]
                group_key = item.get("invoice_no") or f"__row_{row['row']}"
                groups.setdefault(group_key, []).append(item)
            for lines_for_invoice in groups.values():
                first = lines_for_invoice[0]
                customer_id = customers_map.get(_clean(first["customer_name"]).lower())
                invoice_date = _parse_date(first["date"])
                if not customer_id or not invoice_date:
                    skipped += len(lines_for_invoice)
                    continue
                line_payload = []
                ok = True
                for entry in lines_for_invoice:
                    product = products_map.get(_clean(entry["product"]).lower())
                    if not product:
                        ok = False
                        break
                    product_id, _buy_price = product
                    quantity = float(entry.get("quantity") or 0)
                    unit_price = float(entry.get("unit_price") or 0)
                    discount = float(entry.get("discount") or 0)
                    line_total = max(0.0, quantity * unit_price - discount)
                    line_payload.append((product_id, quantity, unit_price, line_total))
                if not ok or not line_payload:
                    skipped += len(lines_for_invoice)
                    continue
                total_amount = round(sum(item[3] for item in line_payload), 2)
                paid_amount = min(float(first.get("paid_amount") or 0), total_amount)
                payment_status = calculate_payment_status(total_amount, paid_amount, policy["decimal_places"], policy["rounding_mode"])
                created_at_dt = datetime.combine(invoice_date, datetime.min.time())
                result = conn.execute(text("""
                    INSERT INTO invoices
                      (invoice_type, customer_id, subtotal, discount_percent, discount_amount,
                       tax_percent, tax_amount, shipping_cost, total_amount, payment_status,
                       status, payment_terms_days, due_date, invoice_note, qr_enabled, created_at, company_id)
                    VALUES
                      (:invoice_type, :customer_id, :subtotal, 0, 0,
                       0, 0, 0, :total_amount, :payment_status,
                       'final', 0, '', :note, 0, :created_at, :company_id)
                """), {
                    "invoice_type": first["invoice_type"], "customer_id": customer_id,
                    "subtotal": total_amount, "total_amount": total_amount,
                    "payment_status": payment_status,
                    "note": "Imported historical invoice",
                    "created_at": created_at_dt,
                    "company_id": company_id,
                })
                invoice_id = result.lastrowid
                for product_id, quantity, unit_price, line_total in line_payload:
                    conn.execute(text("""
                        INSERT INTO invoice_items (invoice_id, product_id, quantity, unit_price, total_price, company_id)
                        VALUES (:invoice_id, :product_id, :quantity, :unit_price, :total_price, :company_id)
                    """), {
                        "invoice_id": invoice_id, "product_id": product_id, "quantity": quantity,
                        "unit_price": unit_price, "total_price": line_total,
                        "company_id": company_id,
                    })
                _post_historical_invoice(
                    conn, invoice_id, customer_id, first["invoice_type"],
                    total_amount, paid_amount, created_at_dt, policy, company_id,
                )
                inserted += len(lines_for_invoice)

        result = {"inserted": inserted, "updated": updated, "skipped": skipped, "applied_by": actor}
        conn.execute(text("""
            UPDATE data_import_batches
            SET status='applied', result_json=:result, applied_at=:applied_at
            WHERE id=:id
        """), {"result": json.dumps(result), "applied_at": _now(), "id": batch_id})
        return {"status": "applied", "batch_id": batch_id, **result}


@router.post("/rollback/{batch_id}")
def rollback_import_batch(batch_id: str, request: Request):
    actor = _require_admin(request)
    company_id = current_company_id(request)
    with engine.begin() as conn:
        _ensure_schema(conn)
        batch = conn.execute(text(
            "SELECT * FROM data_import_batches WHERE id=:id AND company_id=:company_id"
        ), {"id": batch_id, "company_id": company_id}).mappings().first()
        if not batch:
            raise HTTPException(status_code=404, detail="Import batch not found")
        if batch["entity_type"] != "products":
            raise HTTPException(status_code=400, detail="Rollback is only supported for product import batches")
        if batch["status"] != "applied":
            raise HTTPException(status_code=409, detail=f"Batch cannot be rolled back from status '{batch['status']}'")

        product_ids = [
            row[0] for row in conn.execute(text("""
                SELECT product_id FROM data_import_product_links WHERE batch_id=:batch_id
            """), {"batch_id": batch_id}).fetchall()
        ]
        if not product_ids:
            raise HTTPException(status_code=409, detail="No newly-created products are linked to this batch - nothing to roll back")

        note = f"Import batch {batch_id}"
        removed, skipped_report = [], []
        for product_id in product_ids:
            product = conn.execute(text("SELECT id, name FROM products WHERE id=:id"), {"id": product_id}).mappings().first()
            if not product:
                continue
            invoice_ref = conn.execute(text(
                "SELECT id FROM invoice_items WHERE product_id=:id LIMIT 1"
            ), {"id": product_id}).first()
            if invoice_ref:
                skipped_report.append({"product_id": product_id, "name": product["name"], "reason": "Used on an invoice since import"})
                continue
            other_movement = conn.execute(text("""
                SELECT id FROM stock_movements
                WHERE product_id=:id AND NOT (movement_type='opening_import' AND note=:note)
                LIMIT 1
            """), {"id": product_id, "note": note}).first()
            if other_movement:
                skipped_report.append({"product_id": product_id, "name": product["name"], "reason": "Stock adjusted after import"})
                continue

            movement = conn.execute(text("""
                SELECT warehouse_id, quantity FROM stock_movements
                WHERE product_id=:id AND movement_type='opening_import' AND note=:note
            """), {"id": product_id, "note": note}).mappings().first()
            if movement and movement["warehouse_id"]:
                orm_session = OrmSession(bind=conn)
                apply_warehouse_delta(orm_session, movement["warehouse_id"], product_id, -float(movement["quantity"] or 0), company_id)
                orm_session.flush()
            conn.execute(text("""
                DELETE FROM stock_movements WHERE product_id=:id AND movement_type='opening_import' AND note=:note
            """), {"id": product_id, "note": note})
            try:
                delete_source_voucher("product_opening", product_id, company_id, connection=conn)
            except ValueError as error:
                skipped_report.append({"product_id": product_id, "name": product["name"], "reason": str(error)})
                continue
            conn.execute(text("DELETE FROM products WHERE id=:id"), {"id": product_id})
            conn.execute(text(
                "DELETE FROM data_import_product_links WHERE batch_id=:batch_id AND product_id=:id"
            ), {"batch_id": batch_id, "id": product_id})
            removed.append({"product_id": product_id, "name": product["name"]})

        new_status = "rolled_back" if not skipped_report else "partially_rolled_back"
        result = {"removed": removed, "skipped": skipped_report, "rolled_back_by": actor}
        conn.execute(text("""
            UPDATE data_import_batches SET status=:status, result_json=:result WHERE id=:id
        """), {"status": new_status, "result": json.dumps(result, ensure_ascii=False), "id": batch_id})
        return {"status": new_status, "batch_id": batch_id, **result}


@router.get("/batches")
def list_import_batches(request: Request):
    _require_preview_role(request, entity="products")
    with engine.begin() as conn:
        _ensure_schema(conn)
        rows = conn.execute(text("""
            SELECT id, entity_type, file_name, status, total_rows, valid_rows,
                   error_rows, duplicate_rows, result_json, created_by, created_at, applied_at
            FROM data_import_batches WHERE company_id=:company_id ORDER BY created_at DESC LIMIT 100
        """), {"company_id": current_company_id(request)}).mappings().all()
        return [dict(row) for row in rows]


@router.delete("/batches/{batch_id}")
def delete_import_batch(batch_id: str, request: Request):
    """Admin-only cleanup of an import batch's stored history/preview data
    (the parsed row payload, kept for audit/report purposes, can grow large
    for big files). Only ever deletes the history record itself - never the
    products/customers/etc. it already created; use /rollback for that."""
    _require_admin(request)
    with engine.begin() as conn:
        _ensure_schema(conn)
        batch = conn.execute(text(
            "SELECT id, status FROM data_import_batches WHERE id=:id AND company_id=:company_id"
        ), {"id": batch_id, "company_id": current_company_id(request)}).mappings().first()
        if not batch:
            raise HTTPException(status_code=404, detail="Import batch not found")
        if batch["status"] == "applied":
            raise HTTPException(
                status_code=409,
                detail="Applied batches must be rolled back (or left as an audit record) before their history can be deleted",
            )
        conn.execute(text("DELETE FROM data_import_batches WHERE id=:id"), {"id": batch_id})
        conn.execute(text("DELETE FROM data_import_product_links WHERE batch_id=:id"), {"id": batch_id})
        return {"status": "deleted", "batch_id": batch_id}
