from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

EXCEL_PATH = "vetrix_invoices.xlsx"


def _get(obj, key, default=""):
    if obj is None:
        return default
    if isinstance(obj, dict):
        return obj.get(key, default)
    return getattr(obj, key, default)


def _customer_name(customers, customer_id):
    customer = (customers or {}).get(customer_id)
    if customer is None:
        customer = (customers or {}).get(str(customer_id))
    if isinstance(customer, dict):
        return customer.get("name") or customer.get("customer_name") or str(customer_id or "")
    return str(customer if customer is not None else customer_id or "")


def build_invoice_excel(invoices, settings=None, customers=None, language="en", filename=None):
    fa = language == "fa"
    decimal_places = max(0, min(4, int(_get(settings, "decimal_places", 2) or 0)))
    currency_code = _get(settings, "currency_code", "") or _get(settings, "currency", "") or ""
    wb = Workbook()
    ws = wb.active
    ws.title = "فاکتورها" if fa else "Invoices"
    ws.sheet_view.rightToLeft = fa
    headers = (
        ["شماره", "تاریخ", "نوع", "طرف‌حساب", "جمع جزء", "تخفیف", "مالیات", "حمل", "مبلغ نهایی", "تسویه‌شده", "باقی‌مانده", "وضعیت", "ارز"]
        if fa else
        ["ID", "Date", "Type", "Party", "Subtotal", "Discount", "Tax", "Shipping", "Grand total", "Settled", "Remaining", "Status", "Currency"]
    )
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="0F172A")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    number_format = "#,##0" + (("." + ("0" * decimal_places)) if decimal_places else "")
    for inv in invoices:
        total = float(_get(inv, "total_amount", 0) or 0)
        settled = float(_get(inv, "settled_amount", 0) or 0)
        remaining = float(_get(inv, "remaining_amount", total - settled) or 0)
        ws.append([
            _get(inv, "id", ""),
            str(_get(inv, "created_at", "") or "")[:19].replace("T", " "),
            _get(inv, "invoice_type", ""),
            _customer_name(customers, _get(inv, "customer_id", "")),
            float(_get(inv, "subtotal", 0) or 0),
            float(_get(inv, "discount_amount", 0) or 0),
            float(_get(inv, "tax_amount", 0) or 0),
            float(_get(inv, "shipping_cost", 0) or 0),
            total,
            settled,
            remaining,
            _get(inv, "payment_status", _get(inv, "status", "")),
            currency_code,
        ])
        for column in range(5, 12):
            ws.cell(ws.max_row, column).number_format = number_format

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = [10, 21, 16, 28, 16, 16, 16, 16, 18, 18, 18, 16, 12]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="right" if fa else "left", vertical="center")

    output = filename or EXCEL_PATH
    wb.save(output)
    return output
