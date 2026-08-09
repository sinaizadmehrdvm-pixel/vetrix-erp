"""XLSX builder for the Customers/Parties list export. Mirrors the styling
conventions already established in app/export/excel_export.py (styled
header row, freeze panes, auto-filter, per-column widths, right/left cell
alignment by language) rather than inventing a second convention.
"""
import os
import tempfile
import uuid

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.export.security import sanitize_cell_value

HEADER_FILL = PatternFill("solid", fgColor="0F172A")


def build_customers_list_excel(rows, headers, numeric_columns, meta, language="en", filename=None):
    """rows: list[list] already formatted values (numeric columns stay real
    numbers so Excel keeps them as numbers, not text).
    numeric_columns: set of 0-based column indexes that should get a
    thousands-separated number format instead of plain text alignment.
    """
    fa = language == "fa"
    wb = Workbook()
    ws = wb.active
    ws.title = meta.get("sheet_title") or ("طرف‌حساب‌ها" if fa else "Customers")
    ws.sheet_view.rightToLeft = fa

    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        ws.append([sanitize_cell_value(cell) for cell in row])

    for row in ws.iter_rows(min_row=2):
        for index, cell in enumerate(row):
            if index in numeric_columns:
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="right" if fa else "left", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = meta.get("column_widths") or [14] * len(headers)
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width

    output = filename or os.path.join(tempfile.gettempdir(), f"vetrix_customers_{uuid.uuid4().hex}.xlsx")
    wb.save(output)
    return output
