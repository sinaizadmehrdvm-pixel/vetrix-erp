"""SECURITY PHASE E regression tests for two info-leak fixes:

1. sanitize_cell_value (app/export/security.py) is now applied in
   app/export/excel_export.py (invoice XLSX export) and
   app/report_delivery.py's generate_csv - previously only
   app/export/customers_excel.py guarded against CSV/XLSX formula
   injection, leaving these two export paths open to it.
2. app/sms_utils.py::send_sms and app/telegram_utils.py's send functions
   now convert a raw httpx error (whose message embeds the provider's
   API key/bot token directly in the request URL) into a sanitized
   ValueError before it can reach a caller that persists/displays
   str(error) - e.g. purchase_orders.py's dispatch log, later served back
   verbatim via GET /{po_id}/dispatch-log.
"""
import os
import tempfile
import unittest
from unittest.mock import patch

import httpx
from openpyxl import load_workbook

from app.export.security import sanitize_cell_value


class SanitizeCellValueTests(unittest.TestCase):
    def test_formula_trigger_characters_are_neutralized(self):
        for trigger in ("=", "+", "-", "@", "\t", "\r"):
            value = f"{trigger}cmd|'/c calc'!A1"
            self.assertEqual(sanitize_cell_value(value), f"'{value}")

    def test_ordinary_text_is_untouched(self):
        self.assertEqual(sanitize_cell_value("Regular Customer Name"), "Regular Customer Name")

    def test_non_string_values_pass_through(self):
        self.assertEqual(sanitize_cell_value(42), 42)
        self.assertEqual(sanitize_cell_value(3.14), 3.14)
        self.assertIsNone(sanitize_cell_value(None))


class InvoiceExcelExportSanitizationTests(unittest.TestCase):
    def test_malicious_customer_name_is_neutralized_in_the_exported_cell(self):
        from app.export.excel_export import build_invoice_excel

        malicious_name = "=cmd|'/c calc'!A1"
        invoices = [{
            "id": 1, "created_at": "2026-01-01T00:00:00", "invoice_type": "sale",
            "customer_id": 1, "total_amount": 1000, "settled_amount": 0,
            "subtotal": 1000, "discount_amount": 0, "tax_amount": 0, "shipping_cost": 0,
            "payment_status": "pending",
        }]
        # Explicit filename - build_invoice_excel defaults to a fixed,
        # tracked path (backend/vetrix_invoices.xlsx) when none is given.
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = os.path.join(temp_dir, "test_output.xlsx")
            path = build_invoice_excel(
                invoices, customers={1: {"name": malicious_name}}, language="en", filename=temp_path,
            )
            wb = load_workbook(path)
            ws = wb.active
            party_cell = ws.cell(row=2, column=4).value  # "Party" column
            self.assertTrue(str(party_cell).startswith("'"), party_cell)
            self.assertIn(malicious_name, party_cell)


class ReportDeliveryCsvSanitizationTests(unittest.TestCase):
    def test_malicious_row_value_is_neutralized_in_generated_csv(self):
        from app import report_delivery as report_delivery_module

        malicious_row = {"name": "=cmd|'/c calc'!A1", "amount": 1000}
        with patch.object(
            report_delivery_module, "_rows_for",
            return_value=({"columns": [("name", "Name"), ("amount", "Amount")]}, [malicious_row]),
        ):
            csv_bytes = report_delivery_module.generate_csv("any", company_id=1)
        text = csv_bytes.decode("utf-8-sig")
        self.assertIn("'=cmd", text)
        self.assertNotIn("\n=cmd", text)


class ProviderCredentialLeakTests(unittest.TestCase):
    def test_sms_http_status_error_does_not_leak_the_api_key(self):
        from app import sms_utils

        secret_key = "SUPER-SECRET-KAVENEGAR-KEY-12345"
        request = httpx.Request("POST", f"https://api.kavenegar.com/v1/{secret_key}/sms/send.json")
        response = httpx.Response(401, request=request, text="unauthorized")

        with patch.object(sms_utils, "sms_settings", return_value={"panel": "kavenegar", "api_key": secret_key}):
            with patch.object(httpx, "post", return_value=response):
                with self.assertRaises(ValueError) as ctx:
                    sms_utils.send_sms(1, "09120000000", "test message")
        self.assertNotIn(secret_key, str(ctx.exception))
        self.assertIn("401", str(ctx.exception))

    def test_sms_connection_error_does_not_leak_the_api_key(self):
        from app import sms_utils

        secret_key = "SUPER-SECRET-KAVENEGAR-KEY-67890"

        def _raise_connect_error(*args, **kwargs):
            raise httpx.ConnectError(f"Connection failed for url https://api.kavenegar.com/v1/{secret_key}/sms/send.json")

        with patch.object(sms_utils, "sms_settings", return_value={"panel": "kavenegar", "api_key": secret_key}):
            with patch.object(httpx, "post", side_effect=_raise_connect_error):
                with self.assertRaises(ValueError) as ctx:
                    sms_utils.send_sms(1, "09120000000", "test message")
        self.assertNotIn(secret_key, str(ctx.exception))

    def test_telegram_send_message_connection_error_does_not_leak_the_bot_token(self):
        from app import telegram_utils

        secret_token = "123456:SUPER-SECRET-BOT-TOKEN"

        def _raise_connect_error(*args, **kwargs):
            raise httpx.ConnectError(f"Connection failed for url https://api.telegram.org/bot{secret_token}/sendMessage")

        with patch.object(telegram_utils, "telegram_bot_token", return_value=secret_token):
            with patch.object(httpx, "post", side_effect=_raise_connect_error):
                with self.assertRaises(ValueError) as ctx:
                    telegram_utils.send_telegram_message(1, "12345", "test message")
        self.assertNotIn(secret_token, str(ctx.exception))

    def test_telegram_send_document_connection_error_does_not_leak_the_bot_token(self):
        from app import telegram_utils

        secret_token = "123456:SUPER-SECRET-BOT-TOKEN-2"

        def _raise_connect_error(*args, **kwargs):
            raise httpx.ConnectError(f"Connection failed for url https://api.telegram.org/bot{secret_token}/sendDocument")

        with patch.object(telegram_utils, "telegram_bot_token", return_value=secret_token):
            with patch.object(httpx, "post", side_effect=_raise_connect_error):
                with self.assertRaises(ValueError) as ctx:
                    telegram_utils.send_telegram_document(1, "12345", "file.txt", b"content")
        self.assertNotIn(secret_token, str(ctx.exception))


if __name__ == "__main__":
    unittest.main()
